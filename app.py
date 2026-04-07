"""
app.py
------
Aplicação Flask — interface web para geração de listas PF.

Rotas:
  GET  /              → formulário de filtros
  POST /levantamento  → contagem rápida antes de gerar (sem download)
  POST /gerar         → gera lista, retorna preview + link de download
  GET  /download      → baixa o Excel gerado
  GET  /bairros/<cidade> → retorna JSON com bairros da cidade (IBGE + OSM)
"""

import os
import re
import time
from datetime import datetime

import mysql.connector
import pandas as pd
from flask import (Flask, jsonify, redirect, render_template,
                   request, send_file, session, url_for)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from bairros_api import cidades_para_select, obter_bairros
from config import DB_CONFIG, DIR_OUTPUT, LOG_LEVEL
from data_cleaner import limpar_dataframe, relatorio_html
from data_processor import colunas_saida, processar
from list_logger import configurar_logging, registrar_erro, registrar_geracao
from query_builder import build_query, descrever_filtros_db

# ── Logging estruturado ──────────────────────────────────────────
log = configurar_logging(LOG_LEVEL)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", os.urandom(24))
# Para produção: defina a variável de ambiente FLASK_SECRET com valor fixo.

UFS_BRASIL = [
    "AC","AL","AM","AP","BA","CE","DF","ES","GO","MA",
    "MG","MS","MT","PA","PB","PE","PI","PR","RJ","RN",
    "RO","RR","RS","SC","SE","SP","TO",
]


# ============================================================
# UTILITÁRIOS
# ============================================================

def conectar_banco():
    return mysql.connector.connect(**DB_CONFIG)


def gerar_excel_formatado(df: pd.DataFrame, filepath: str) -> None:
    """
    Gera arquivo Excel com formatações baseadas na macro VBA.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista PF"
    
    # Escrever cabeçalhos
    headers = df.columns.tolist()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Escrever dados
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Formatações específicas por coluna
            col_name = headers[col_num - 1]
            
            if col_name in ['NOME', 'CPF']:
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
            elif col_name.startswith('DDD_'):
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
                cell.number_format = '00'
            elif col_name.startswith('TELEFONE_'):
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
                cell.number_format = '000000000'  # 9 dígitos
            elif col_name in ['GENERO', 'UF']:
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
            elif col_name == 'DATA_NASCIMENTO':
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
            elif col_name in ['ENDERECO', 'BAIRRO', 'CIDADE']:
                cell.alignment = Alignment(horizontal='left', vertical='bottom')
            elif col_name.startswith('EMAIL_'):
                cell.alignment = Alignment(horizontal='left', vertical='bottom')
            elif col_name == 'CBO':
                cell.alignment = Alignment(horizontal='left', vertical='bottom')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='bottom')
    
    # Ajustar largura das colunas
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].auto_size = True
    
    # Zoom
    ws.sheet_view.zoomScale = 90
    
    wb.save(filepath)


def executar_query(sql: str, params: list) -> pd.DataFrame:
    conn = conectar_banco()
    try:
        df = pd.read_sql(sql, conn, params=params)
    finally:
        conn.close()
    return df


def extrair_filtros(form) -> dict:
    """
    Lê o formulário HTML e devolve dicionário de filtros.
    Separadores aceitos em campos de lista: vírgula, ponto-e-vírgula, nova linha.
    """
    def lista(campo):
        raw = form.get(campo, "")
        if not raw:
            return []
        return [i.strip() for i in re.split(r"[,;\n]+", raw) if i.strip()]

    def inteiro(campo):
        val = form.get(campo, "").strip()
        return int(val) if val.lstrip("-").isdigit() else None

    def decimal(campo):
        val = form.get(campo, "").strip().replace(",", ".")
        try:
            return float(val)
        except ValueError:
            return None

    return {
        # ── Filtros que vão ao BANCO ─────────────────────────────
        "ufs":           lista("ufs"),
        "cidades":       lista("cidades"),
        "bairros":       lista("bairros"),
        "genero":        form.get("genero", "ambos"),
        "idade_min":     inteiro("idade_min"),
        "idade_max":     inteiro("idade_max"),
        "email":         form.get("email", "nao_filtrar"),
        # ── Filtros que ficam no PYTHON ──────────────────────────
        "tipo_telefone": form.get("tipo_telefone", "movel"),
        "cbos":          lista("cbos"),
        "quantidade":    inteiro("quantidade"),
    }


# ============================================================
# ROTAS
# ============================================================

@app.route("/")
def index():
    return render_template(
        "index.html",
        ufs=UFS_BRASIL,
        cidades_nobres=cidades_para_select(),
    )


@app.route("/bairros/<cidade>")
def bairros_da_cidade(cidade: str):
    """
    API JSON: consulta bairros da cidade via IBGE + OpenStreetMap (Overpass).
    Resultados ficam em cache por 24h.

    Retorna:
      { cidade, bairros: [...], total, fonte, erro? }
    """
    try:
        bairros = obter_bairros(cidade)
        return jsonify({
            "cidade": cidade.upper(),
            "bairros": bairros,
            "total": len(bairros),
            "fonte": "OpenStreetMap / IBGE Localidades",
        })
    except Exception as exc:
        app.logger.exception("Erro ao buscar bairros de '%s'", cidade)
        return jsonify({
            "cidade": cidade.upper(),
            "bairros": [],
            "total": 0,
            "erro": str(exc),
        })


@app.route("/levantamento", methods=["POST"])
def levantamento():
    """
    Levantamento rápido: retorna contagem de registros no banco
    com os filtros indexados (UF, Cidade, Bairro, Gênero, Idade, Email).
    Permite avaliar o volume antes de gerar a lista completa.
    """
    try:
        filtros = extrair_filtros(request.form)
        sql, params = build_query(filtros)

        # Substitui o SELECT por COUNT(*)
        sql_count = "SELECT COUNT(*) AS total\n" + sql[sql.index("FROM"):]

        conn = conectar_banco()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(sql_count, params)
        resultado = cursor.fetchone()
        cursor.close()
        conn.close()

        total = resultado.get("total", 0) if resultado else 0
        log.info("Levantamento: %d registros | %s", total, descrever_filtros_db(filtros))
        return jsonify({
            "ok": True,
            "total_banco": total,
            "descricao": descrever_filtros_db(filtros),
            "aviso": (
                "Este total ainda será reduzido pelos filtros Python: "
                "tipo de telefone e CBO."
            ) if total > 0 else None,
        })
    except Exception as e:
        log.warning("Erro no levantamento: %s", e)
        return jsonify({"ok": False, "erro": str(e)}), 400


@app.route("/gerar", methods=["POST"])
def gerar():
    """
    Fluxo principal de geração de lista:
      1. Monta query (filtros indexados → banco)
      2. Executa no banco → DataFrame bruto
      3. Limpeza de sujeiras (data_cleaner) — rastreado para auditoria
      4. Aplica filtros Python (telefone, CBO, quantidade)
      5. Gera Excel (.xlsx) com formatações aplicadas em output/
      6. Registra geração no log de auditoria (logs/geracoes/geracoes.csv)
      7. Retorna preview HTML + link de download
    """
    t0 = time.perf_counter()
    filtros = {}
    try:
        filtros = extrair_filtros(request.form)
        sql, params = build_query(filtros)
        log.info("GERAR iniciado | %s", descrever_filtros_db(filtros))
        log.debug("SQL: %s | params: %s", sql, params)

        df_bruto    = executar_query(sql, params)
        total_banco = len(df_bruto)
        log.info("Banco retornou %d registros.", total_banco)

        # ── Limpeza para auditoria (contagem pré-filtros Python) ──
        df_limpo_audit, _rel_dict = limpar_dataframe(df_bruto.copy())
        total_apos_limpeza = len(df_limpo_audit)

        # ── Pipeline completo (limpeza + filtros Python) ──────────
        df_filtrado, rel_limpeza = processar(df_bruto, filtros)
        total_final = len(df_filtrado)
        log.info("Após filtros Python: %d registros.", total_final)

        if total_final == 0:
            duracao = time.perf_counter() - t0
            registrar_geracao(filtros, total_banco, total_apos_limpeza,
                              0, "", duracao, status="VAZIO")
            return render_template(
                "index.html",
                ufs=UFS_BRASIL,
                cidades_nobres=cidades_para_select(),
                erro="Nenhum registro encontrado com os filtros informados.",
                rel_limpeza=rel_limpeza,
                filtros=filtros,
            )

        # ── Salvar Excel em output/ ─────────────────────────────────
        incluir_email = filtros.get("email") != "nao_filtrar"
        cols_desejadas  = colunas_saida(com_email=incluir_email)
        cols_existentes = [c for c in cols_desejadas if c in df_filtrado.columns]
        df_saida = df_filtrado[cols_existentes].copy()

        ufs_str      = "_".join(filtros.get("ufs", ["geral"]))
        ts           = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"lista_pf_{ufs_str}_{ts}.xlsx"
        caminho      = DIR_OUTPUT / nome_arquivo
        gerar_excel_formatado(df_saida, str(caminho))

        duracao = time.perf_counter() - t0
        log.info("Lista gerada: %s | %d registros | %.1fs",
                 nome_arquivo, total_final, duracao)

        # ── Auditoria ─────────────────────────────────────────────
        registrar_geracao(
            filtros=filtros,
            total_banco=total_banco,
            total_apos_limpeza=total_apos_limpeza,
            total_final=total_final,
            nome_arquivo=nome_arquivo,
            duracao_s=duracao,
            status="OK",
        )

        session["arquivo_gerado"] = str(caminho)
        session["nome_arquivo"]   = nome_arquivo

        return render_template(
            "index.html",
            ufs=UFS_BRASIL,
            cidades_nobres=cidades_para_select(),
            preview=df_saida.head(50).to_dict(orient="records"),
            colunas_preview=list(df_saida.columns),
            total_banco=total_banco,
            total_final=total_final,
            nome_arquivo=nome_arquivo,
            rel_limpeza=rel_limpeza,
            filtros=filtros,
        )

    except ValueError as ve:
        duracao = time.perf_counter() - t0
        registrar_erro(filtros, str(ve), duracao)
        log.warning("Filtros inválidos: %s", ve)
        return render_template(
            "index.html", ufs=UFS_BRASIL,
            cidades_nobres=cidades_para_select(),
            erro=str(ve), filtros=filtros or {}
        )
    except Exception as e:
        duracao = time.perf_counter() - t0
        registrar_erro(filtros, str(e), duracao)
        log.exception("Erro ao gerar lista")
        return render_template(
            "index.html", ufs=UFS_BRASIL,
            cidades_nobres=cidades_para_select(),
            erro=f"Erro interno: {e}", filtros=filtros or {}
        )


@app.route("/download")
def download():
    caminho = session.get("arquivo_gerado")
    nome    = session.get("nome_arquivo", "lista.xlsx")
    if not caminho or not os.path.exists(caminho):
        return redirect(url_for("index"))
    return send_file(caminho, 
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nome)


# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
